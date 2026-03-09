import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# where to find the CSVs and where to save the output
bom_folder  = Path("BOMs")
output_file = Path("BOM_Standardisation.xlsx")

# background colours for each component class, used across all sheets
class_colours = {
    "Capacitor":"FFF2CC",
    "Resistor":"E2EFDA",
    "Diode":"FCE4D6",
    "LED":"E8D5F5",
    "IC":"DDEEFF",
    "Inductor":"FDE9D9",
    "Connector":"D9EAD3",
    "Switch":"F4CCCC",
    "Oscillator": "CFE2F3",
    "Transistor": "FFE5CC",
    "Other":"FFFFFF"
}

# shared style constants
row_alt= "EEF2FB"  # alternating row colour
col_totals = "D6E4F0" # background for total/summary columns
grey = "AAAAAA" # used for zero/dash entries
navy = "1A2C5B"# main dark blue used for headers and bold text

thin_line = Side(style="thin", color="CCCCCC")
cell_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

# navy header cell with white bold text
def style_header(cell, text):
    cell.value = text
    cell.font = Font(name="Aptos", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=navy, fgColor=navy)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

# standard data cell, bg and font colour are optional
def style_cell(cell, value, bg="FFFFFF", bold=False, center=False, color="000000"):
    cell.value= value
    cell.font = Font(name="Aptos", bold=bold, color=color, size=10)
    cell.fill = PatternFill("solid", start_color=bg, fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = cell_border

# works out what type of component something is
# checks the designator prefix since descriptions from altium are inconsistent
def classify(description, designator):
    des = str(designator).upper()
    if des.startswith("DS"):return "LED"
    if des.startswith("C"): return "Capacitor"
    if des.startswith("R"): return "Resistor"
    if des.startswith("D"): return "Diode"
    if des.startswith("L"): return "Inductor"
    if des.startswith("J"): return "Connector"
    if des.startswith("SW"): return "Switch"
    if des.startswith("Y") or des.startswith("X"): return "Oscillator"
    if des.startswith("Q"): return "Transistor"
    if des.startswith("U"): return "IC"
    return "Other"

# load every CSV in the BOMs folder
csv_files = [f.name for f in bom_folder.glob("*.csv")]

all_boards = []
for file in csv_files:
    csv_path = bom_folder / file
    board_df = pd.read_csv(csv_path)

    # altium exports have a blank first column and a bunch of supplier columns we don't need
    board_df = board_df.drop(columns=board_df.columns[0])
    board_df = board_df.drop(columns=["Supplier Unit Price 1", "Supplier Subtotal 1", "Revision ID","Revision Status", "Manufacturer 1", "Manufacturer Lifecycle 1",
                                       "Supplier 1", "Supplier Part Number 1"], errors="ignore")
    board_df = board_df.rename(columns={
        "Line #":"line",
        "Quantity":"qty",
        "Name":"name",
        "Description":"description",
        "Manufacturer Part Number 1":"part_number",
        "Designator":"designator",
    })
    board_df["class"] = board_df.apply(lambda row: classify(row["description"], row["designator"]), axis=1)
    board_df["board"] = csv_path.stem  # use filename as board name
    all_boards.append(board_df)

# combine all boards into one dataframe, sorted by board > class > name
df = pd.concat(all_boards, ignore_index=True)
df = df.sort_values(["board", "class", "name"]).reset_index(drop=True)

# keep board order consistent with how files were loaded
board_names = list(dict.fromkeys(df["board"]))

wb = openpyxl.Workbook()

# --Master Component List--
# one row per unique component, one column per board showing quantity
ws_master = wb.active
ws_master.title = "Master Component List"

# pivot so each board becomes its own column
pivot = df.pivot_table(
    index=["name", "description", "class", "part_number"],
    columns="board",
    values="qty",
    aggfunc="sum"
).fillna(0).reset_index()

pivot = pivot.sort_values(["class", "name"]).reset_index(drop=True)

# reindex forces all boards to appear even if they share no components with others
pivot = pivot.reindex(columns=["name", "description", "class", "part_number"] + board_names, fill_value=0)
pivot[board_names] = pivot[board_names].astype(int)

# column position helpers
board_col_start = 5
board_col_end = 4 + len(board_names)
total_qty_col = board_col_end + 1
boards_used_col = board_col_end + 2

headers = ["Name", "Description", "Class", "Part Number"] + board_names + ["Total Qty", "Boards Used"]
for col, header in enumerate(headers, 1):
    style_header(ws_master.cell(row=1, column=col), header)
ws_master.row_dimensions[1].height = 32

for row_idx, (_, row) in enumerate(pivot.iterrows(), 2):
    bg = row_alt if row_idx % 2 == 0 else "FFFFFF"

    style_cell(ws_master.cell(row=row_idx, column=1), row["name"], bg=bg, bold=True)
    style_cell(ws_master.cell(row=row_idx, column=2), row["description"], bg=bg)

    cls_bg = class_colours.get(row["class"], "FFFFFF")
    style_cell(ws_master.cell(row=row_idx, column=3), row["class"], bg=cls_bg, bold=True, center=True)

    style_cell(ws_master.cell(row=row_idx, column=4), row["part_number"], bg=bg)

    for i, board in enumerate(board_names):
        qty = int(row[board])
        cell = ws_master.cell(row=row_idx, column=board_col_start + i)
        if qty == 0:
            style_cell(cell, "-", bg=bg, center=True, color=grey)
        else:
            style_cell(cell, qty, bg=bg, bold=True, center=True, color=navy)

    # excel formulas so totals update if you edit quantities directly in the sheet
    qty_range = f"{get_column_letter(board_col_start)}{row_idx}:{get_column_letter(board_col_end)}{row_idx}"

    tot_cell = ws_master.cell(row=row_idx, column=total_qty_col)
    tot_cell.value = f'=SUMIF({qty_range},">0")'
    tot_cell.font= Font(name="Aptos", bold=True, color=navy, size=10)
    tot_cell.fill  = PatternFill("solid", start_color=col_totals, fgColor=col_totals)
    tot_cell.alignment = Alignment(horizontal="center", vertical="center")
    tot_cell.border = cell_border

    bu_cell = ws_master.cell(row=row_idx, column=boards_used_col)
    bu_cell.value = f'=COUNTIF({qty_range},">0")'
    bu_cell.font  = Font(name="Aptos", bold=True, color=navy, size=10)
    bu_cell.fill = PatternFill("solid", start_color=col_totals, fgColor=col_totals)
    bu_cell.alignment = Alignment(horizontal="center", vertical="center")
    bu_cell.border = cell_border

ws_master.auto_filter.ref = f"A1:{get_column_letter(boards_used_col)}{len(pivot) + 1}"
ws_master.freeze_panes = "E2"  # keeps name/desc/class/partnumber visible when scrolling right

ws_master.column_dimensions["A"].width = 20
ws_master.column_dimensions["B"].width = 48
ws_master.column_dimensions["C"].width = 13
ws_master.column_dimensions["D"].width = 22
for i in range(len(board_names)):
    ws_master.column_dimensions[get_column_letter(board_col_start + i)].width = 13
ws_master.column_dimensions[get_column_letter(total_qty_col)].width  = 11
ws_master.column_dimensions[get_column_letter(boards_used_col)].width = 13

# --Raw Data--
# every component from every board in one flat list
# class and description are VLOOKUPs from the master sheet so edits there flow through here
ws_raw = wb.create_sheet("Raw Data")

raw_headers = ["Board", "Class", "Name", "Description", "Qty", "Designator", "Part Number"]
for col, header in enumerate(raw_headers, 1):
    style_header(ws_raw.cell(row=1, column=col), header)
ws_raw.row_dimensions[1].height = 28

for row_idx, (_, row) in enumerate(df.iterrows(), 2):
    bg = row_alt if row_idx % 2 == 0 else "FFFFFF"
    style_cell(ws_raw.cell(row=row_idx, column=1), row["board"], bg=bg, bold=True)

    class_cell = ws_raw.cell(row=row_idx, column=2)
    class_cell.value= f"=IFERROR(VLOOKUP(C{row_idx},'Master Component List'!A:C,3,FALSE),\"Other\")"
    class_cell.font= Font(name="Aptos", bold=True, color=navy, size=10)
    class_cell.fill = PatternFill("solid", start_color=class_colours.get(row["class"], "FFFFFF"),
                                        fgColor=class_colours.get(row["class"], "FFFFFF"))
    class_cell.alignment = Alignment(horizontal="center", vertical="center")
    class_cell.border = cell_border

    style_cell(ws_raw.cell(row=row_idx, column=3), row["name"], bg=bg)

    desc_cell = ws_raw.cell(row=row_idx, column=4)
    desc_cell.value= f"=IFERROR(VLOOKUP(C{row_idx},'Master Component List'!A:B,2,FALSE),\"\")"
    desc_cell.font= Font(name="Aptos", color="000000", size=10)
    desc_cell.fill = PatternFill("solid", start_color=bg, fgColor=bg)
    desc_cell.alignment = Alignment(horizontal="left", vertical="center")
    desc_cell.border = cell_border

    style_cell(ws_raw.cell(row=row_idx, column=5), row["qty"], bg=bg, center=True, bold=True, color=navy)
    style_cell(ws_raw.cell(row=row_idx, column=6), row["designator"], bg=bg)
    style_cell(ws_raw.cell(row=row_idx, column=7), row["part_number"], bg=bg)

ws_raw.auto_filter.ref = f"A1:G{len(df) + 1}"
ws_raw.freeze_panes = "A2"
ws_raw.column_dimensions["A"].width = 16
ws_raw.column_dimensions["B"].width = 13
ws_raw.column_dimensions["C"].width = 20
ws_raw.column_dimensions["D"].width = 48
ws_raw.column_dimensions["E"].width = 8
ws_raw.column_dimensions["F"].width = 30
ws_raw.column_dimensions["G"].width = 20

# --Board Summary--
# one row per board, shows total components and a breakdown by class
ws_summary = wb.create_sheet("Board Summary")

class_list = ["Capacitor", "Resistor", "Diode", "LED", "IC", "Inductor", "Connector", "Switch", "Oscillator", "Transistor", "Other"]
last_col = 3 + len(class_list)

ws_summary.merge_cells(f"A1:{get_column_letter(last_col)}1")
title = ws_summary["A1"]
title.value = "BOARD SUMMARY"
title.font = Font(name="Aptos", bold=True, size=14, color="FFFFFF")
title.fill = PatternFill("solid", start_color=navy, fgColor=navy)
title.alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 30

summary_headers = ["Board", "Total Components", "Unique Components"] + class_list
for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=2, column=col, value=header)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border
    if col <= 3:
        cell.font = Font(name="Aptos", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color=navy, fgColor=navy)
    else:
        cls_bg = class_colours.get(class_list[col - 4], "FFFFFF")
        cell.font = Font(name="Aptos", bold=True, color=navy, size=10)
        cell.fill = PatternFill("solid", start_color=cls_bg, fgColor=cls_bg)
ws_summary.row_dimensions[2].height = 32

for row_idx, board in enumerate(board_names, 3):
    board_df = df[df["board"] == board]
    total_qty = int(board_df["qty"].sum())
    unique = len(board_df)
    bg = row_alt if row_idx % 2 == 0 else "FFFFFF"

    style_cell(ws_summary.cell(row=row_idx, column=1), board, bg=bg, bold=True)
    style_cell(ws_summary.cell(row=row_idx, column=2), total_qty, bg=col_totals, bold=True, center=True, color=navy)
    style_cell(ws_summary.cell(row=row_idx, column=3), unique,bg=col_totals, bold=True, center=True, color=navy)

    for i, cls in enumerate(class_list):
        qty   = int(board_df[board_df["class"] == cls]["qty"].sum())
        cls_bg = class_colours.get(cls, "FFFFFF")
        cell  = ws_summary.cell(row=row_idx, column=4 + i)
        if qty == 0:
            style_cell(cell, "", bg=cls_bg, center=True)
        else:
            style_cell(cell, qty, bg=cls_bg, bold=True, center=True, color=navy)

ws_summary.auto_filter.ref = f"A2:{get_column_letter(last_col)}{1 + len(board_names) + 1}"
ws_summary.freeze_panes = "A3"
ws_summary.column_dimensions["A"].width = 18
ws_summary.column_dimensions["B"].width = 18
ws_summary.column_dimensions["C"].width = 18
for i in range(len(class_list)):
    ws_summary.column_dimensions[get_column_letter(4 + i)].width = 13

wb.save(output_file)
print(f"done, saved to {output_file}")